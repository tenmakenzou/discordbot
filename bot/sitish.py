import openpyxl
from datetime import datetime, timedelta

def run(day):
    wb = openpyxl.load_workbook('menu.xlsx')
    
    # Pick target date
    if day == "today":
        target_date = datetime.now()
    elif day == "tomorrow":
        target_date = datetime.now() + timedelta(days=1)
    else:
        return "Invalid day — use 'today' or 'tomorrow'."

    # Find the cell that matches the date
    def get_row(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    if cell.value.date() == target_date.date():
                        return cell.coordinate
        return None

    x = None
    active_sheet = None
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        x = get_row(sheet)
        if x:
            active_sheet = sheet
            break

    if not x:
        return f"No schedule found for {day} ({target_date.date()})"

    

    # Words to ignore (headers, weekdays, empty cells)
    skip_list = [
        None, "", " ", "Γάλα φρέσκο ζεστό ή κρύο, Τσάι σε διάφορες γεύσεις, μαρμελάδες σε διάφορες γεύσεις, Μέλι, Φρυγανιές σίτου, Ψωμί, Κέικ, βούτυρο ή ταχίνι, αυγό",
        'Δευτέρα', 'Τρίτη', 'Τετάρτη', 'Πέμπτη', 'Παρασκευή', 'Σαββάτο', 'Κυριακή',
        'Δευτέρα-Monday', 'Τρίτη-Tuesday', 'Τετάρτη-Wednesday', 'Πέμπτη-Thursday',
        'Παρασκευή-Friday', 'Σάββατο-Saturday', 'Κυριακή-Sunday',"ή"
    ]

    def should_skip(value):
        if (value in skip_list) or (isinstance(value, str) and "ΕΒΔΟΜΑΔΑ" in value.upper()):
            return True
        return False

    def get_date(sheet, coord):
        count = 0
        result = ""

        col = coord[0]
        row_num = int(coord[1:])

        for i in range(row_num, sheet.max_row + 1):
            cell = sheet[f"{col}{i}"]
            if not isinstance(cell.value, datetime) and not should_skip(cell.value):
                if count == 0:
                    result += '\n***Πρωινό***\n- ' + "Γάλα φρέσκο ζεστό ή κρύο, Τσάι σε διάφορες γεύσεις, μαρμελάδες σε διάφορες γεύσεις, Μέλι, Φρυγανιές σίτου, Ψωμί, Κέικ, βούτυρο ή ταχίνι, αυγό" + '\n\n***Μεσημεριανό ***\n'
                if count == 4:
                    result += "\n***Δείπνο ***\n"
                result += f"- {cell.value}\n"
                count += 1
                if count == 7:
                    break
        return result.strip()

    return get_date(active_sheet, x)
